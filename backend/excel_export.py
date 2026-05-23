"""Excel export for support tickets.

Appends each created ticket as a row to backend/exports/support_tickets.xlsx
so the support team can triage offline. Creates the file with a styled
header on first use. Failure to write is logged but never breaks the API
response — the ticket is in the DB regardless.
"""
import logging
import threading
from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

log = logging.getLogger("savomart.excel")

EXPORT_DIR = Path(__file__).resolve().parent / "exports"
EXPORT_PATH = EXPORT_DIR / "support_tickets.xlsx"
SHEET_NAME = "Tickets"

HEADERS = [
    "Ticket ID",
    "Timestamp (UTC)",
    "User ID",
    "Mobile",
    "Name",
    "Category",
    "Subject",
    "Description",
    "Source",
    "Status",
]

# openpyxl is not thread-safe for the same file handle. Multiple uvicorn
# workers would each have their own lock — for a single-process dev/demo
# setup this is sufficient; for production with multiple workers we'd put
# this behind a queue or a real DB-backed export job.
_lock = threading.Lock()


def _init_workbook() -> Workbook:
    wb = Workbook()
    ws = wb.active
    ws.title = SHEET_NAME
    ws.append(HEADERS)

    header_fill = PatternFill(start_color="782B90", end_color="782B90", fill_type="solid")
    header_font = Font(color="FFF200", bold=True)
    center = Alignment(horizontal="left", vertical="center")
    for col_idx, _ in enumerate(HEADERS, start=1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center

    widths = [14, 22, 8, 14, 22, 18, 38, 60, 8, 12]
    for col_idx, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = w

    ws.freeze_panes = "A2"
    return wb


def append_ticket_row(row: dict[str, Any]) -> None:
    """Append one ticket row. Safe to call from request handlers."""
    try:
        with _lock:
            EXPORT_DIR.mkdir(parents=True, exist_ok=True)
            if EXPORT_PATH.exists():
                wb = load_workbook(EXPORT_PATH)
                ws = wb[SHEET_NAME] if SHEET_NAME in wb.sheetnames else wb.active
            else:
                wb = _init_workbook()
                ws = wb[SHEET_NAME]

            ts = row.get("timestamp")
            if isinstance(ts, datetime):
                ts = ts.strftime("%Y-%m-%d %H:%M:%S")

            ws.append([
                row.get("ticket_id", ""),
                ts or "",
                row.get("user_id", ""),
                row.get("mobile", ""),
                row.get("name", ""),
                row.get("category", ""),
                row.get("subject", ""),
                row.get("description", ""),
                row.get("source", ""),
                row.get("status", ""),
            ])
            wb.save(EXPORT_PATH)
    except Exception as exc:  # noqa: BLE001 — never fail the request
        log.warning("excel export failed: %s", exc)
