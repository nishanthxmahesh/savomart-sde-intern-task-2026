"""Excel import for customer activity.

Accepts an .xlsx with the columns (case-insensitive):
  name, mobile, amount_spent, coupon_code, description

Loyalty rules engine:
  - Every ₹10 of `amount_spent` earns 1 loyalty point (floor division).
  - `coupon_code`, if provided, looks up the user's matching active coupon
    and marks it as used (records a 0-delta REDEMPTION transaction so it
    appears in their ledger).
  - Both can happen in the same row (bought ₹500 of stuff using SAVE10).

User handling:
  - Lookup by 10-digit mobile.
  - Exists → adjust points + tier.
  - Missing → create a new Bronze customer with `name`, points start at 0,
    then apply the row's points.
"""
from __future__ import annotations

import io
import logging
from dataclasses import dataclass, field
from datetime import datetime, timezone
from typing import Optional

from openpyxl import load_workbook
from sqlalchemy.orm import Session

from loyalty import tier_for_balance
from models import Coupon, PointsSource, PointsTransaction, Tier, User

log = logging.getLogger("savomart.import")

POINTS_PER_RUPEE = 1 / 10  # ₹10 = 1 point

EXPECTED_COLUMNS = ["name", "mobile", "amount_spent", "coupon_code", "description"]


@dataclass
class RowOutcome:
    row: int
    mobile: str
    name: str
    points_awarded: int = 0
    coupon_redeemed: Optional[str] = None
    created: bool = False
    error: Optional[str] = None


@dataclass
class ImportResult:
    total_rows: int = 0
    created_users: int = 0
    updated_users: int = 0
    points_awarded_total: int = 0
    coupons_redeemed_total: int = 0
    outcomes: list[RowOutcome] = field(default_factory=list)
    errors: list[RowOutcome] = field(default_factory=list)


def _clean_mobile(raw) -> Optional[str]:
    if raw is None:
        return None
    s = str(raw).strip().replace(" ", "").replace("-", "")
    if s.startswith("+91"):
        s = s[3:]
    if s.startswith("91") and len(s) == 12:
        s = s[2:]
    # Excel often turns numbers into "9999999999.0"
    if s.endswith(".0"):
        s = s[:-2]
    return s if s.isdigit() and len(s) == 10 else None


def _to_int(raw) -> int:
    if raw is None or raw == "":
        return 0
    try:
        # Accept floats (Excel default) and strings
        return int(float(raw))
    except (TypeError, ValueError):
        return 0


def _to_str(raw) -> str:
    return str(raw).strip() if raw is not None else ""


def _read_header(ws) -> dict[str, int]:
    """Return a {column_name_lower: 0-based-col-index} mapping from row 1."""
    mapping: dict[str, int] = {}
    for idx, cell in enumerate(ws[1]):
        key = _to_str(cell.value).lower().replace(" ", "_")
        if key:
            mapping[key] = idx
    return mapping


def parse_workbook(content: bytes) -> tuple[list[dict], list[str]]:
    """Open the xlsx, validate headers, return (rows, header_errors).

    Each row is a dict of normalized values. Empty rows are skipped.
    """
    errors: list[str] = []
    try:
        wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    except Exception as exc:  # noqa: BLE001
        return [], [f"Couldn't open the file. Is it a valid .xlsx? ({exc})"]

    ws = wb.active
    if ws is None or ws.max_row < 1:
        return [], ["The workbook is empty."]

    header = _read_header(ws)
    missing_required = [c for c in ("name", "mobile") if c not in header]
    if missing_required:
        errors.append(
            f"Missing required column(s): {', '.join(missing_required)}. "
            f"Expected headers: {', '.join(EXPECTED_COLUMNS)}."
        )
        return [], errors

    rows: list[dict] = []
    for r_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if not any(c is not None and str(c).strip() != "" for c in row):
            continue  # skip blank lines

        def col(name: str):
            i = header.get(name)
            return row[i] if i is not None and i < len(row) else None

        rows.append({
            "row": r_idx,
            "name": _to_str(col("name")),
            "mobile_raw": col("mobile"),
            "amount_spent": _to_int(col("amount_spent")),
            "coupon_code": _to_str(col("coupon_code")).upper(),
            "description": _to_str(col("description")),
        })

    return rows, errors


def apply_rows(db: Session, rows: list[dict], admin_id: Optional[int] = None) -> ImportResult:
    """Apply each parsed row. Commits in one go at the end — caller commits."""
    result = ImportResult(total_rows=len(rows))
    now = datetime.now(timezone.utc)

    for r in rows:
        outcome = RowOutcome(row=r["row"], mobile=str(r["mobile_raw"] or ""), name=r["name"])

        mobile = _clean_mobile(r["mobile_raw"])
        if not mobile:
            outcome.error = "Invalid mobile (need 10 digits)"
            result.errors.append(outcome)
            continue

        if not r["name"] or len(r["name"]) < 2:
            outcome.error = "Name must be at least 2 characters"
            result.errors.append(outcome)
            continue

        amount = max(0, int(r["amount_spent"] or 0))
        points = int(amount * POINTS_PER_RUPEE)  # floor via int cast on /10
        coupon_code = r["coupon_code"] or None

        user = db.query(User).filter(User.mobile_number == mobile).first()
        if user is None:
            user = User(
                name=r["name"][:120],
                mobile_number=mobile,
                points_balance=0,
                tier=Tier.BRONZE,
                is_active=True,
            )
            db.add(user)
            db.flush()  # get user.id
            outcome.created = True
            result.created_users += 1
        else:
            result.updated_users += 1

        outcome.mobile = mobile
        outcome.name = user.name

        # 1. Award points for the purchase
        if points > 0:
            user.points_balance += points
            user.tier = tier_for_balance(user.points_balance)
            desc = f"+{points} pts from ₹{amount} purchase"
            if r["description"]:
                desc += f" — {r['description'][:120]}"
            db.add(PointsTransaction(
                user_id=user.id,
                delta=points,
                description=desc,
                source=PointsSource.PURCHASE,
                admin_id=admin_id,
            ))
            outcome.points_awarded = points
            result.points_awarded_total += points

        # 2. Redeem coupon if provided
        if coupon_code:
            coupon = (
                db.query(Coupon)
                .filter(
                    Coupon.user_id == user.id,
                    Coupon.code == coupon_code,
                    Coupon.is_used == False,  # noqa: E712
                )
                .first()
            )
            if coupon is None:
                outcome.error = (
                    f"Coupon {coupon_code} not found on this user (or already used). "
                    "Purchase points were still applied."
                ) if points > 0 else f"Coupon {coupon_code} not found on this user (or already used)."
                if points == 0:
                    result.errors.append(outcome)
                    continue
            else:
                coupon.is_used = True
                db.add(PointsTransaction(
                    user_id=user.id,
                    delta=0,
                    description=f"Redeemed coupon {coupon_code}",
                    source=PointsSource.REDEMPTION,
                    admin_id=admin_id,
                ))
                outcome.coupon_redeemed = coupon_code
                result.coupons_redeemed_total += 1

        user.last_login_at = user.last_login_at or now  # touch so the timeline updates
        result.outcomes.append(outcome)

    return result


def build_template_workbook() -> bytes:
    """Return a small example .xlsx as bytes (5 columns + 3 example rows)."""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill

    wb = Workbook()
    ws = wb.active
    ws.title = "Customers"

    headers = EXPECTED_COLUMNS
    ws.append(headers)

    for col_idx, _ in enumerate(headers, start=1):
        c = ws.cell(row=1, column=col_idx)
        c.font = Font(bold=True, color="FFFFFFFF")
        c.fill = PatternFill("solid", fgColor="FF782B90")
        c.alignment = Alignment(horizontal="left", vertical="center")

    ws.append(["Aanya Sharma", "9999999999", 1500, "", "Grocery run — Amul + Britannia"])
    ws.append(["Rahul Mehta", "8888888888", 800, "WELCOME10", "Bought with welcome coupon"])
    ws.append(["New Customer", "6123456789", 500, "", "First visit — auto-enrolled"])

    widths = [22, 14, 14, 14, 36]
    from openpyxl.utils import get_column_letter
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.freeze_panes = "A2"

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
